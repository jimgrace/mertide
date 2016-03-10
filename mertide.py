# pip install openpyxl

import openpyxl
import random
import string
from xml.sax.saxutils import escape

inDir = 'input/'
comboDir = 'input/catCombos/'
outDir = 'output/'

def uid():
    uid = random.choice(string.ascii_letters);
    for i in range(0, 10):
        uid += random.choice(string.ascii_letters+'0123456789');
    return uid

random.seed()
wb = openpyxl.load_workbook('mertide.xlsx', read_only = True, data_only = True)
tabs = wb.get_sheet_names()
toc = wb.get_sheet_by_name('TOC')
export = open(outDir+'merforms.xml', 'w')
export.write(open(inDir+'prefix.xml').read())
for i in range(2, toc.max_row + 1):
    tab = toc.cell(row=i, column=1).value
    if tab in tabs:
        datasetCode = str(toc.cell(row=i, column=2).value).strip()
        datasetShortname = str(toc.cell(row=i, column=3).value).strip()
        datasetName = str(toc.cell(row=i, column=4).value).strip()
        sheet = wb.get_sheet_by_name(tab)
        print('Processing ' + datasetName)
        export.write(open(inDir+'dataset_prefix.xml').read().format(datasetCode, datasetName, datasetShortname, uid()))
        form = open(inDir+'form_prefix.html').read()
        form += '<div id="PEPFAR_Tabs_vertical">\n' + \
                '<ul>\n'
        vtab_body = ''
        content = ''
        inHtab = False
        inVtab = False
        inIndicator = False
        for j in range(2, sheet.max_row + 2):
            if (j > sheet.max_row):
                type = 'END'
            else:
                type = str(sheet.cell(row=j, column=1).value).strip()
                x = str(sheet.cell(row=j, column=2).value).strip()
                code = str(sheet.cell(row=j, column=3).value).strip()
                shortname = str(sheet.cell(row=j, column=4).value).strip()
                name = str(sheet.cell(row=j, column=5).value).strip()
                details = str(sheet.cell(row=j, column=6).value).strip()
            if (inIndicator & (type in ['Indicator', 'HTAB', 'VTAB', 'END'])):
                content += open(inDir+'indicator_suffix.html').read()
                inIndicator = False;
            if (inHtab & (type in ['HTAB', 'VTAB', 'END'])):
                content += '</div>\n\n'
                inHtab = False;
            if (inVtab & (type in ['VTAB', 'END'])):
                vtab_body += '</ul>\n\n' +\
                             content
                content = ''
                inVtab = False;
            if type == 'VTAB':
                form += '\t<li><a href="#PEPFAR_Tabs_vertical_' + code + '">' + shortname + '</a></li>\n'
                vtab_body += '\n<div id="PEPFAR_Tabs_vertical_' + code + '">\n' +\
                             '<ul>\n'
                vtab_code = code
                inVtab = True
            elif type == 'HTAB':
                vtab_body += '\t<li><a href="#PEPFAR_Form_' + vtab_code + '_' + code + '">' + shortname + '</a></li>\n'
                content += '<!-- HTAB Close -->\n<div id="PEPFAR_Form_' + vtab_code + '_' + code + '">\n'
                inHtab = True
            elif type == 'Indicator':
                content += open(inDir+'indicator_prefix.html').read().format(shortname, name)
                inIndicator = True
            elif type == 'Required':
                content += open(inDir+'required.html').read().format(name)
            elif type == 'Conditional':
                content += open(inDir+'conditional.html').read().format(name)
            elif type == 'Numerator':
                content += open(inDir+'numerator.html').read().format(uid()) #TODO: Replace with real UID
            elif type == 'DE':
                 content += open(comboDir+code.replace('/','-')+'.html').read().format(uid()) #TODO: Replace with real Data Element and Disagg.
            elif type == 'Subtotal':
                content += open(inDir+'subtotal.html').read().format(uid()) #TODO: Replace with real UID
            elif type != "END":
                print("Unrecognized type " + type + " on line " + str(j) + " of sheet " + tab )
        form += '</ul>\n' +\
                vtab_body +\
                '</div>\n'
        formFile = open(outDir+datasetName+'.html', 'w')
        formFile.write('<html>\n<head>\n' +\
                       '<script src="https://ajax.googleapis.com/ajax/libs/jquery/1.8.2/jquery.min.js"></script>\n' + \
                       '<link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/font-awesome/4.2.0/css/font-awesome.min.css">\n' + \
                       '<link rel="stylesheet" href="https://ajax.googleapis.com/ajax/libs/jqueryui/1.10.4/themes/smoothness/jquery-ui.css">\n' + \
                       '<script type="text/javascript" src="https://dhis2-cdn.org/v215/ext/ext-all.js"></script>\n' + \
                       '<script src="https://ajax.googleapis.com/ajax/libs/jqueryui/1.10.4/jquery-ui.min.js"></script>\n' + \
                       '</head>\n<body>\n' + form + '\n</body>\n</html>\n')
        formFile.close()
        export.write(escape(form))
        export.write('			</htmlCode>\n' + \
                     '		</dataset>\n')
export.write('	</datasets>\n' + \
             '</metadata>\n') #TODO: Write subtotal indicators we have created
export.close()
